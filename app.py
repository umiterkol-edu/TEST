import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, date
from io import BytesIO
import numpy as np
from openpyxl.styles import Font, PatternFill

st.set_page_config(page_title="Katsayı Hesaplama", layout="centered")
st.title("📊 Puan Hesaplama Aracı")

if 'veriler' not in st.session_state:
    st.session_state['veriler'] = []
if 'edit_index' not in st.session_state:
    st.session_state['edit_index'] = None

st.subheader("Yeni Kayıt Ekle")

sehir = st.text_input("İl", placeholder="Örn: Antalya")

col1, col2 = st.columns(2)
with col1:
    baslangic = st.date_input("Başlangıç Tarihi", value=date.today(), format="DD.MM.YYYY", min_value=date(2000, 1, 1))
with col2:
    bitis = st.date_input("Bitiş Tarihi", value=date.today(), format="DD.MM.YYYY", min_value=date(2000, 1, 1))

col3, col4 = st.columns(2)
with col3:
    rapor = st.number_input("Raporlu Gün", min_value=0, step=1)
with col4:
    katsayi = st.number_input("Katsayı", min_value=0.0, step=0.001, format="%.3f")

# Kayıt Ekle/Düzenle
if st.session_state['edit_index'] is None:
    if st.button("✅ Kaydet"):
        yeni = {
            'Şehir': sehir,
            'Başlangıç': baslangic.strftime('%d.%m.%Y'),
            'Bitiş': bitis.strftime('%d.%m.%Y'),
            'Rapor': rapor,
            'Katsayı': katsayi
        }
        st.session_state['veriler'].append(yeni)
        st.success(f"{sehir} eklendi.")
else:
    if st.button("💾 Güncelle"):
        index = st.session_state['edit_index']
        st.session_state['veriler'][index] = {
            'Şehir': sehir,
            'Başlangıç': baslangic.strftime('%d.%m.%Y'),
            'Bitiş': bitis.strftime('%d.%m.%Y'),
            'Rapor': rapor,
            'Katsayı': katsayi
        }
        st.session_state['edit_index'] = None
        st.success("Kayıt güncellendi.")

st.markdown("---")

# Veriler varsa göster
if st.session_state['veriler']:
    st.subheader("📋 Mevcut Kayıtlar")

    for i, item in enumerate(st.session_state['veriler']):
        col1, col2, col3 = st.columns([4, 1, 1])
        with col1:
            st.markdown(f"**{item['Şehir']}** | {item['Başlangıç']} - {item['Bitiş']} | Rapor: {item['Rapor']} gün | Katsayı: {item['Katsayı']}")
        with col2:
            if st.button("✏️ Düzenle", key=f"edit_{i}"):
                st.session_state['edit_index'] = i
                sehir = item['Şehir']
                baslangic = datetime.strptime(item['Başlangıç'], "%d.%m.%Y").date()
                bitis = datetime.strptime(item['Bitiş'], "%d.%m.%Y").date()
                rapor = item['Rapor']
                katsayi = item['Katsayı']
                st.rerun()
        with col3:
            if st.button("🗑️ Sil", key=f"sil_{i}"):
                st.session_state['veriler'].pop(i)
                st.success("Kayıt silindi.")
                st.experimental_rerun()

    st.markdown("---")
    results = []
    toplam_sonuc = 0
    for item in st.session_state['veriler']:
        start = datetime.strptime(item['Başlangıç'], '%d.%m.%Y')
        end = datetime.strptime(item['Bitiş'], '%d.%m.%Y')
        toplam_gun = (end - start).days
        net_gun = toplam_gun - item['Rapor']
        sonuc = net_gun * item['Katsayı']
        toplam_sonuc += sonuc
        results.append({
            'Şehir': item['Şehir'],
            'Başlangıç': item['Başlangıç'],
            'Bitiş': item['Bitiş'],
            'Toplam Gün': toplam_gun,
            'Rapor': item['Rapor'],
            'Net Gün': net_gun,
            'Katsayı': item['Katsayı'],
            'Sonuç': round(sonuc, 3)
        })

    df = pd.DataFrame(results)
    df.loc[len(df)] = ['Toplam', np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, round(toplam_sonuc, 3)]

    st.subheader("🧮 Hesaplama Tablosu")
    st.dataframe(df, use_container_width=True)

    st.subheader("📈 Grafik")
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(df['Şehir'][:-1], df['Sonuç'][:-1], color='skyblue')
    ax.set_xlabel("Şehir")
    ax.set_ylabel("Sonuç")
    ax.set_title("Şehirlere Göre Sonuç")
    ax.grid(True, linestyle="--", alpha=0.7)
    st.pyplot(fig)

    st.subheader("📥 Dosya İndir")

    excel_io = BytesIO()
    with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hesap')
        ws = writer.book['Hesap']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        total_fill = PatternFill("solid", fgColor="D9E1F2")
        for cell in ws[ws.max_row]:
            cell.fill = total_fill

    st.download_button("📊 Excel İndir", data=excel_io.getvalue(),
                       file_name="katsayi_hesaplama.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    pdf_io = BytesIO()
    fig.savefig(pdf_io, format='pdf')
    st.download_button("🖼️ Grafik PDF İndir", data=pdf_io.getvalue(),
                       file_name="sonuc_grafik.pdf", mime="application/pdf")

# Tüm kayıtları temizle
if st.button("🧹 Tüm Kayıtları Temizle"):
    st.session_state['veriler'] = []
    st.success("Tüm kayıtlar silindi.")
